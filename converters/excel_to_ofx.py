from io import BytesIO
from datetime import datetime
from openpyxl import load_workbook
from converters.ofx_builder import build_ofx


def convert_excel_to_ofx(file_bytes, bank_id="000", account_id="000000", account_type="CHECKING"):
    workbook = load_workbook(BytesIO(file_bytes), data_only=True)
    sheet = workbook.active
    rows = list(sheet.iter_rows(values_only=True))

    transactions = []

    for row in rows[1:]:
        if not row:
            continue

        date = row[0] if len(row) > 0 else None
        memo = row[3] if len(row) > 3 else ""
        amount = row[4] if len(row) > 4 else None

        # Layout alternativo simples: Data | Descrição | Valor
        if not isinstance(amount, (int, float)) and len(row) > 2:
            alt_amount = row[2]
            alt_memo = row[1] if len(row) > 1 else ""
            if isinstance(alt_amount, (int, float)):
                memo = alt_memo
                amount = alt_amount

        if isinstance(date, datetime) and isinstance(amount, (int, float)):
            transactions.append({
                "date": date,
                "memo": memo or "",
                "amount": float(amount)
            })

    if not transactions:
        raise ValueError("Nenhuma transação válida encontrada no Excel.")

    ofx = build_ofx(
        transactions=transactions,
        bank_id=bank_id,
        account_id=account_id,
        account_type=account_type
    )

    return ofx, len(transactions)
