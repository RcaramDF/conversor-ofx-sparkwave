from io import BytesIO
from datetime import datetime
from openpyxl import load_workbook
from converters.ofx_builder import build_ofx


def parse_brazilian_money(value):
    if isinstance(value, (int, float)):
        return float(value)

    value = str(value).replace("R$", "").replace(" ", "").strip()
    negative = "-" in value
    value = value.replace("-", "")
    value = value.replace(".", "").replace(",", ".")

    try:
        amount = float(value)
        return -amount if negative else amount
    except Exception:
        return None


def convert_excel_to_ofx(file_bytes, bank_id="000", account_id="000000", account_type="CHECKING"):
    workbook = load_workbook(BytesIO(file_bytes), data_only=True)
    sheet = workbook.active
    rows = list(sheet.iter_rows(values_only=True))

    transactions = []

    for row in rows[1:]:
        if not row:
            continue

        date = row[0] if len(row) > 0 else None
        memo = row[1] if len(row) > 1 and row[1] else "Movimentação"
        amount_raw = row[2] if len(row) > 2 else None

        if len(row) > 4 and row[4] is not None:
            memo = row[3] if row[3] else memo
            amount_raw = row[4]

        amount = parse_brazilian_money(amount_raw)

        if isinstance(date, datetime) and amount is not None:
            transactions.append({
                "date": date,
                "memo": memo,
                "amount": amount
            })

    if not transactions:
        raise ValueError("Nenhuma transação válida encontrada no Excel.")

    ofx = build_ofx(
        transactions=transactions,
        bank_id=bank_id,
        account_id=account_id,
        account_type=account_type
    )

    return ofx, len(transactions)